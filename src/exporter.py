"""Excel workbook export module using Pandas and OpenPyXL for multi-sheet reporting."""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.settings import (
    DEFAULT_OUTPUT_DIR,
    DEMAND_INDICATOR_IDS,
    EXCEL_BORDER_COLOR,
    EXCEL_PRIMARY_FILL_COLOR,
    EXCEL_PRIMARY_SECTION_KEYWORDS,
    EXCEL_SECONDARY_FILL_COLOR,
    EXCEL_SECTION_GAP,
    PRICE_INDICATOR_IDS
)
from src.utils import translate_indicator, translate_geography

# OpenPyXL Style Objects initialized from centralized settings
PRIMARY_FILL = PatternFill(start_color=EXCEL_PRIMARY_FILL_COLOR, end_color=EXCEL_PRIMARY_FILL_COLOR, fill_type="solid")
PRIMARY_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

SECONDARY_FILL = PatternFill(start_color=EXCEL_SECONDARY_FILL_COLOR, end_color=EXCEL_SECONDARY_FILL_COLOR, fill_type="solid")
SECONDARY_FONT = Font(name="Calibri", size=11, bold=True, color=EXCEL_PRIMARY_FILL_COLOR)

THIN_BORDER = Border(
    left=Side(style="thin", color=EXCEL_BORDER_COLOR),
    right=Side(style="thin", color=EXCEL_BORDER_COLOR),
    top=Side(style="thin", color=EXCEL_BORDER_COLOR),
    bottom=Side(style="thin", color=EXCEL_BORDER_COLOR),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


def _write_section(
    writer: pd.ExcelWriter,
    sheet_name: str,
    title: str,
    df_data: pd.DataFrame,
    start_row: int,
) -> int:
    """Writes a standardized section consisting of a section title row merged across
    data columns and a Pandas DataFrame.

    Args:
        writer (pd.ExcelWriter): OpenPyXL Excel writer instance.
        sheet_name (str): Target worksheet name.
        title (str): Section header title text.
        df_data (pd.DataFrame): Data table to write into the worksheet.
        start_row (int): 1-based row index to begin writing the section title.

    Returns:
        int: Next available 1-based row index after inserting spacing gap.
    """
    # Write Data Table first (creates the worksheet if it doesn't exist)
    data_start_row = start_row + 1
    df_data.to_excel(writer, sheet_name=sheet_name, startrow=data_start_row - 1, index=False)

    ws = writer.sheets[sheet_name]

    # Determine table width and merge title cells across all columns
    num_cols = len(df_data.columns)
    if num_cols > 1:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_cols)

    # Write Section Title in top-left cell of the merged region
    ws.cell(row=start_row, column=1, value=title)

    # Calculate next section starting row including gap
    next_row = start_row + 1 + len(df_data) + EXCEL_SECTION_GAP
    return next_row


def _apply_workbook_styles(file_path: Path) -> None:
    """Applies professional styling, color hierarchies, thin borders,
    and auto-fits column widths accurately across all worksheets.

    Args:
        file_path (Path): File path to the Excel workbook.

    Raises:
        Exception: Logged if loading or saving the workbook fails.
    """
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print(f"⚠️ Could not load workbook for styling: {e}")
        return

    for sheet in wb.worksheets:
        sheet.views.sheetView[0].showGridLines = True
        column_header_rows = set()

        # Identify section title rows and column header rows
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            first_cell_val = row[0].value
            if isinstance(first_cell_val, str) and any(kw in first_cell_val for kw in EXCEL_PRIMARY_SECTION_KEYWORDS):
                column_header_rows.add(row_idx + 1)

        # Apply styles strictly within exact table bounds
        for row in sheet.iter_rows():
            first_cell_val = row[0].value
            row_idx = row[0].row

            is_primary_row = isinstance(first_cell_val, str) and any(kw in first_cell_val for kw in EXCEL_PRIMARY_SECTION_KEYWORDS)
            is_secondary_row = row_idx in column_header_rows

            if is_primary_row:
                # Find the exact width (max_column) of the merged range for this specific primary title
                max_title_col = 1
                for merged_range in sheet.merged_cells.ranges:
                    if merged_range.min_row == row_idx:
                        max_title_col = merged_range.max_col
                        break

                # Apply header styling ONLY up to the table's merged width
                for col_idx in range(1, max_title_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.fill = PRIMARY_FILL
                    cell.font = PRIMARY_FONT
                    cell.alignment = ALIGN_CENTER
                    cell.border = THIN_BORDER

            elif is_secondary_row:
                # Count only the actual non-empty columns for this table header
                num_header_cols = len([c for c in row if c.value is not None])

                # Apply secondary styling ONLY up to the table's actual column count
                for col_idx in range(1, num_header_cols + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.fill = SECONDARY_FILL
                    cell.font = SECONDARY_FONT
                    cell.alignment = ALIGN_CENTER
                    cell.border = THIN_BORDER

            else:
                # Standard data rows: format existing cell values only
                for cell in row:
                    if cell.value is not None:
                        cell.border = THIN_BORDER

                        if isinstance(cell.value, float):
                            cell.number_format = "#,##0.00"

        # Step 3: Auto-fit column widths dynamically
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len

            adjusted_width = min(max(max_len + 3, 12), 50)
            sheet.column_dimensions[col_letter].width = adjusted_width

    try:
        wb.save(file_path)
    except Exception as e:
        print(f"⚠️ Could not save styled workbook: {e}")


def _build_executive_summary_tab(
    writer: pd.ExcelWriter,
    min_dt: datetime,
    max_dt: datetime,
    demand_stats: dict,
    price_stats: dict,
    market_volume_stats: dict,
) -> None:
    """Builds the Executive Summary worksheet of the Excel report.

    Args:
        writer (pd.ExcelWriter): Excel writer used to create and populate the workbook.
        min_dt (datetime): Start datetime of the analysis period.
        max_dt (datetime): End datetime of the analysis period.
        demand_stats (dict): Statistical indicators calculated for energy demand.
        price_stats (dict): Statistical indicators calculated for energy prices.
        market_volume_stats (dict): Statistical indicators calculated for market volume and economic activity.
    """
    sheet_name = "Executive Summary"
    current_row = 1

    # Metadata Section
    meta_data = {
        "Metric": ["Analysis Period Start", "Analysis Period End", "Report Generation Time"],
        "Value": [
            min_dt.strftime("%Y-%m-%d %H:%M:%S"),
            max_dt.strftime("%Y-%m-%d %H:%M:%S"),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ],
    }
    current_row = _write_section(writer, sheet_name, "ANALYSIS METADATA & PERIOD", pd.DataFrame(meta_data), current_row)

    # Market Volume Section
    if market_volume_stats:
        volume_data = {
            "Market Indicator": [
                "Total Traded Volume (M€)",
                "Total Energy Processed (GWh)",
                "Volume-Weighted Avg Price (VWAP)",
                "Peak Expenditure Timestamp",
                "Peak Hourly Cost (k€)",
                "Peak Hour Demand (MW)",
                "Peak Hour SPOT Price (€/MWh)",
            ],
            "Value": [
                market_volume_stats.get("total_volume_eur", 0.0) / 1_000_000,
                market_volume_stats.get("total_energy_mwh", 0.0) / 1_000,
                market_volume_stats.get("weighted_avg_price", 0.0),
                str(market_volume_stats.get("peak_spend_hour", "N/A")),
                market_volume_stats.get("peak_spend_eur", 0.0) / 1_000,
                market_volume_stats.get("peak_spend_demand_mw", 0.0),
                market_volume_stats.get("peak_spend_price_eur", 0.0),
            ],
        }
        current_row = _write_section(writer, sheet_name, "MARKET ECONOMIC VOLUME SUMMARY", pd.DataFrame(volume_data), current_row)

    # Performance Overview Section
    summary_rows = [
        {"Indicator": lbl, "Type": "Demand", "Mean": s.get("mean"), "Max": s.get("max"), "Min": s.get("min")}
        for lbl, s in demand_stats.items()
    ] + [
        {"Indicator": lbl, "Type": "Price", "Mean": s.get("mean"), "Max": s.get("max"), "Min": s.get("min")}
        for lbl, s in price_stats.items()
    ]

    if summary_rows:
        _write_section(writer, sheet_name, "KEY INDICATORS PERFORMANCE OVERVIEW", pd.DataFrame(summary_rows), current_row)


def _build_statistical_analysis_tab(writer: pd.ExcelWriter, demand_stats: dict, price_stats: dict) -> None:
    """Builds the Statistical Analysis worksheet of the Excel report.

    Args:
        writer (pd.ExcelWriter): Excel writer used to create and populate the workbook.
        demand_stats (dict): Statistical indicators calculated for energy demand.
        price_stats (dict): Statistical indicators calculated for energy prices.
    """
    sheet_name = "Statistical Analysis"
    current_row = 1

    # Detailed Demand
    demand_rows = [
        {
            "Indicator Name": series_label,
            "Mean (MW)": stats.get("mean"),
            "Median (MW)": stats.get("median"),
            "Std Dev (MW)": stats.get("std_dev"),
            "Max Value (MW)": stats.get("max"),
            "Max Timestamp": str(stats.get("peak_time", "N/A")),
            "Min Value (MW)": stats.get("min"),
        }
        for series_label, stats in demand_stats.items()
    ]
    df_demand = pd.DataFrame(demand_rows)
    if not df_demand.empty:
        current_row = _write_section(writer, sheet_name, "DETAILED DEMAND STATISTICS (MW)", df_demand, current_row)

    # Detailed Prices
    price_rows = [
        {
            "Indicator Name": series_label,
            "Max (€/MWh)": stats.get("max"),
            "Max Timestamp": str(stats.get("max_time", "N/A")),
            "Min (€/MWh)": stats.get("min"),
            "Min Timestamp": str(stats.get("min_time", "N/A")),
            "Spread (€/MWh)": stats.get("spread"),
            "Low Price Hours (<=5€)": stats.get("zero_low_price_hours"),
        }
        for series_label, stats in price_stats.items()
    ]
    df_price = pd.DataFrame(price_rows)
    if not df_price.empty:
        _write_section(writer, sheet_name, "DETAILED PRICE STATISTICS (€/MWh)", df_price, current_row)


def _build_models_anomalies_tab(writer: pd.ExcelWriter, comp_stats: dict, anomalies: dict) -> None:
    """Builds the Models & Anomalies worksheet of the Excel report.

    Args:
        writer (pd.ExcelWriter): Excel writer used to create and populate the workbook.
        comp_stats (dict): Statistics from the pairwise comparison of demand models.
        anomalies (dict): Detected statistical anomalies and outliers grouped by indicator series.
    """
    sheet_name = "Models & Anomalies"
    current_row = 1

    # Model Comparison
    if comp_stats:
        model_a_id = comp_stats.get("model_a")
        model_b_id = comp_stats.get("model_b")

        model_a_name = translate_indicator(indicator_id=model_a_id) if isinstance(model_a_id, int) else str(model_a_id)
        model_b_name = translate_indicator(indicator_id=model_b_id) if isinstance(model_b_id, int) else str(model_b_id)

        df_comp = pd.DataFrame([{
            "Baseline Series": model_a_name,
            "Target Series": model_b_name,
            "MAPE (%)": comp_stats.get("mape"),
            "Pearson Correlation (r)": comp_stats.get("correlation"),
            "Mean Difference (MW)": comp_stats.get("mean_difference"),
            "Max Difference (MW)": comp_stats.get("max_difference_value"),
            "Max Difference Timestamp": str(comp_stats.get("max_difference_time", "N/A")),
        }])
        current_row = _write_section(writer, sheet_name, "PAIRWISE DEMAND MODEL COMPARISON", df_comp, current_row)

    # Anomalies
    anomaly_rows = []
    if isinstance(anomalies, dict):
        for series_label, rec_list in anomalies.items():
            for rec in rec_list:
                anomaly_rows.append({
                    "Timestamp": str(rec.get("datetime")),
                    "Indicator": series_label,
                    "Observed Value (MW)": rec.get("value"),
                    "Deviation": rec.get("deviation"),
                    "Anomaly Type": rec.get("type"),
                })

    df_anomalies = pd.DataFrame(anomaly_rows)
    if not df_anomalies.empty:
        _write_section(writer, sheet_name, "DETECTED STATISTICAL ANOMALIES & OUTLIERS", df_anomalies, current_row)


def _build_clean_data_tab(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
    """Builds the Clean Data worksheet containing processed market data.

    Args:
        writer (pd.ExcelWriter): Excel writer used to create and populate the workbook.
        df (pd.DataFrame): Cleaned and validated energy market data to be exported.
    """
    sheet_name = "Clean Data"
    df_clean = df.copy()

    priority_order = DEMAND_INDICATOR_IDS + PRICE_INDICATOR_IDS

    if "indicator_id" in df_clean.columns:
        df_clean["indicator_id"] = pd.Categorical(df_clean["indicator_id"], categories=priority_order, ordered=True)
        df_clean["name"] = df_clean["indicator_id"].map(lambda x: translate_indicator(indicator_id=x))

    if "geo_id" in df_clean.columns:
        df_clean["geo_name"] = df_clean["geo_id"].map(lambda x: translate_geography(geo_id=x))

    if "datetime" in df_clean.columns:
        df_clean["datetime"] = pd.to_datetime(df_clean["datetime"])

    sort_cols = [col for col in ["indicator_id", "datetime"] if col in df_clean.columns]
    if sort_cols:
        df_clean = df_clean.sort_values(by=sort_cols, ascending=[True, True])

    if "datetime" in df_clean.columns:
        df_clean["datetime"] = df_clean["datetime"].dt.strftime("%Y-%m-%d %H:%M:%S")

    cols_to_keep = ["indicator_id", "name", "geo_id", "geo_name", "datetime", "value"]
    df_export = df_clean[[col for col in cols_to_keep if col in df_clean.columns]]

    _write_section(writer, sheet_name, "CLEAN & PROCESSED MARKET DATA", df_export, start_row=1)


def export_to_excel(
    df: pd.DataFrame,
    demand_stats: dict,
    price_stats: dict,
    comp_stats: dict,
    anomalies: dict,
    market_volume_stats: dict,
    output_dir: str | Path = DEFAULT_OUTPUT_DIR,
) -> str:
    """Generates a styled, multi-tab Excel workbook containing energy market analytics.

    Args:
        df (pd.DataFrame): Cleaned and validated energy market data.
        demand_stats (dict): Statistics for energy demand analysis.
        price_stats (dict): Statistics for energy price analysis.
        comp_stats (dict): Statistics for comparative demand model analysis.
        anomalies (dict): Information about detected statistical anomalies and outliers.
        market_volume_stats (dict): Statistics for market volume and economic analysis.
        output_dir (str | Path): Directory where the generated Excel workbook will be saved.

    Returns:
        str: Full path to the generated Excel workbook.
    """
    print("\n📊 Generating Excel workbook...")

    # Ensure target output directory exists
    out_dir_path = Path(output_dir)
    out_dir_path.mkdir(parents=True, exist_ok=True)

    # Determine dynamic filename based on temporal dataset range
    min_dt, max_dt = df["datetime"].min(), df["datetime"].max()
    fmt = "%Y%m%d_%H%M" if min_dt.date() == max_dt.date() else "%Y%m%d"
    filename = f"energy_analysis_{min_dt.strftime(fmt)}_to_{max_dt.strftime(fmt)}.xlsx"
    file_path = out_dir_path / filename

    # Create Excel Writer
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        _build_executive_summary_tab(writer, min_dt, max_dt, demand_stats, price_stats, market_volume_stats)
        _build_statistical_analysis_tab(writer, demand_stats, price_stats)
        _build_models_anomalies_tab(writer, comp_stats, anomalies)
        _build_clean_data_tab(writer, df)

    # Post-process formatting with OpenPyXL
    _apply_workbook_styles(file_path)

    print(f"✅ Excel report successfully exported to: '{file_path}'")
    return str(file_path)